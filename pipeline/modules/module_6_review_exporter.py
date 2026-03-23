"""Module 6: Exporter — XLSX batch review + HTML cards + ARE skeleton.

Pipeline v4: adapted for the new GeoTask output format with coordinates,
scenario, expected_behavior, and verification fields.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, List, Optional

from shared.config import PipelineConfig
from shared.jsonl_io import read_jsonl

PIPELINE_ROOT = Path(__file__).resolve().parent.parent
DATA_IN = PIPELINE_ROOT / "data" / "05_verified"
DATA_OUT = PIPELINE_ROOT / "data" / "06_export"
INPUT_PATH = DATA_IN / "verified.jsonl"
XLSX_PATH = DATA_OUT / "review_batch.xlsx"
CARDS_DIR = DATA_OUT / "cards"
ARE_DIR = DATA_OUT / "are"


# ═══════════════════════════════════════════════════════════════════════════════
# XLSX export
# ═══════════════════════════════════════════════════════════════════════════════

def _export_xlsx(tasks: List[Dict[str, Any]]) -> Path:
    """Export tasks to XLSX for batch review."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("[Module 6] openpyxl not installed. Skipping XLSX export.")
        return XLSX_PATH

    wb = Workbook()
    ws = wb.active
    ws.title = "Review"

    headers = [
        "image_id", "task_type", "complexity", "image_role",
        "query_preview", "narrative", "must_pass_count",
        "review_priority", "judge_pass", "gt_verified",
    ]
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # Sort by review_priority
    priority_order = {"high": 3, "medium": 2, "low": 1}
    sorted_tasks = sorted(
        tasks,
        key=lambda t: priority_order.get(
            t.get("judge_report", {}).get("review_priority", "low"), 0
        ),
        reverse=True,
    )

    for row_idx, task in enumerate(sorted_tasks, 2):
        coords = task.get("coordinates", {})
        scenario = task.get("scenario", {})
        turns = scenario.get("turns", [])
        verification = task.get("verification", {})
        judge = task.get("judge_report", {})
        gt = task.get("ground_truth", {})

        query = turns[0].get("content", "")[:100] if turns else ""

        values = [
            task.get("provenance", {}).get("image_id", ""),
            coords.get("task_type", ""),
            coords.get("complexity", ""),
            coords.get("image_role", ""),
            query,
            scenario.get("narrative", "")[:80],
            len(verification.get("must_pass", [])),
            judge.get("review_priority", ""),
            judge.get("overall_pass", ""),
            gt.get("verified", ""),
        ]
        for col, v in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=v)

    # Adjust column widths
    widths = [16, 10, 10, 10, 50, 40, 12, 12, 10, 10]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[
            ws.cell(row=1, column=col).column_letter
        ].width = w

    wb.save(XLSX_PATH)
    return XLSX_PATH


# ═══════════════════════════════════════════════════════════════════════════════
# HTML card export
# ═══════════════════════════════════════════════════════════════════════════════

def _export_html_card(task: Dict[str, Any], output_path: Path) -> None:
    """Export a single task as an HTML review card."""
    coords = task.get("coordinates", {})
    scenario = task.get("scenario", {})
    turns = scenario.get("turns", [])
    behavior = task.get("expected_behavior", {})
    verification = task.get("verification", {})
    judge = task.get("judge_report", {})
    gt = task.get("ground_truth", {})
    prov = task.get("provenance", {})
    image_path = prov.get("image_path", "")

    # Use relative path for image
    image_rel = ""
    if image_path:
        try:
            image_rel = str(
                Path(image_path).relative_to(Path(output_path).parent)
            )
        except ValueError:
            image_rel = image_path

    # Build turns HTML
    turns_html = ""
    for t in turns:
        turns_html += (
            f'<div style="background:#fffde7; padding:10px; border-radius:4px; '
            f'margin:5px 0;"><strong>{t.get("role", "user")}:</strong> '
            f'{t.get("content", "")}</div>\n'
        )

    # Build must_pass HTML
    must_pass_html = "\n".join(
        f"<li>{mp}</li>" for mp in verification.get("must_pass", [])
    )

    # Build scores HTML
    scores = judge.get("scores", {})
    scores_html = "\n".join(
        f"<tr><td>{k}</td><td>{v}/5</td></tr>" for k, v in scores.items()
    )

    # Build API calls HTML
    api_html = "\n".join(
        f"<li><code>{c.get('method','')}</code> — {c.get('purpose','')}</li>"
        for c in behavior.get("api_calls", [])
    )

    # GT checks
    gt_checks_html = "\n".join(
        f"<li>{'✓' if c['passed'] else '✗'} {c['check']}: {c.get('detail','')}</li>"
        for c in gt.get("checks", [])
    )

    judge_pass = judge.get("overall_pass", False)

    html = f"""<!DOCTYPE html>
<html lang="zh-CN">
<head><meta charset="UTF-8"><title>Review: {prov.get('image_id','')}</title>
<style>
body {{ font-family: -apple-system, sans-serif; max-width: 900px; margin: 20px auto; padding: 0 15px; }}
.card {{ border: 1px solid #ddd; border-radius: 8px; padding: 20px; margin: 15px 0; }}
.card h3 {{ margin-top: 0; color: #333; }}
img {{ max-width: 100%; border-radius: 6px; }}
table {{ border-collapse: collapse; width: 100%; margin: 10px 0; }}
th, td {{ border: 1px solid #ddd; padding: 6px 10px; text-align: left; }}
th {{ background: #f5f5f5; }}
.tag {{ display: inline-block; background: #e8f0fe; padding: 2px 8px; border-radius: 4px; margin: 2px; font-size: 0.85em; }}
.pass {{ color: green; font-weight: bold; }}
.fail {{ color: red; font-weight: bold; }}
pre {{ background: #f8f8f8; padding: 10px; border-radius: 4px; overflow-x: auto; font-size: 0.85em; }}
</style>
</head>
<body>
<h1>Review Card: {prov.get('image_id','')}</h1>

<div class="card">
<h3>Coordinates</h3>
<span class="tag">task_type: {coords.get('task_type','')}</span>
<span class="tag">complexity: {coords.get('complexity','')}</span>
<span class="tag">image_role: {coords.get('image_role','')}</span>
</div>

<div class="card">
<h3>Image</h3>
<img src="{image_rel}" alt="street view">
<p><strong>Scene:</strong> {prov.get('scene_description', '')[:200]}</p>
</div>

<div class="card">
<h3>Scenario</h3>
<p><em>{scenario.get('narrative', '')}</em></p>
{turns_html}
</div>

<div class="card">
<h3>Expected Behavior</h3>
<p><strong>Reasoning:</strong> {behavior.get('reasoning', '')}</p>
<h4>API Calls</h4>
<ul>{api_html}</ul>
</div>

<div class="card">
<h3>Verification</h3>
<h4>Must Pass</h4>
<ul>{must_pass_html}</ul>
<h4>Answer Reference</h4>
<pre>{json.dumps(verification.get('answer_reference', {}), ensure_ascii=False, indent=2)}</pre>
</div>

<div class="card">
<h3>Judge Report</h3>
<p>Status: <span class="{'pass' if judge_pass else 'fail'}">{'PASS' if judge_pass else 'REJECT'}</span>
   | Priority: {judge.get('review_priority', '')}</p>
<table><tr><th>Dimension</th><th>Score</th></tr>
{scores_html}
</table>
{f'<p><strong>Suggestions:</strong> {judge.get("suggestions", "")}</p>' if judge.get('suggestions') else ''}
</div>

<div class="card">
<h3>Ground Truth Verification</h3>
<p>Status: <span class="{'pass' if gt.get('verified') else 'fail'}">{'VERIFIED' if gt.get('verified') else 'UNVERIFIED'}</span></p>
<ul>{gt_checks_html}</ul>
</div>

</body></html>"""

    output_path.write_text(html, encoding="utf-8")


# ═══════════════════════════════════════════════════════════════════════════════
# ARE skeleton export
# ═══════════════════════════════════════════════════════════════════════════════

_TASK_TYPE_TO_ARE_TAGS = {
    "locate": ["Search", "Execution"],
    "search": ["Search", "Execution"],
    "route": ["Planning", "Execution"],
    "plan": ["Planning", "Time"],
    "judge": ["Ambiguity"],
    "adapt": ["Adaptability", "Time"],
}


def _export_are_skeleton(task: Dict[str, Any]) -> Dict[str, Any]:
    """Convert a pipeline task to an ARE ExportedTrace skeleton.

    This is a best-effort mapping for future ARE integration.
    A full GeoApp does not exist yet, so the skeleton is incomplete.
    """
    prov = task.get("provenance", {})
    coords = task.get("coordinates", {})
    scenario = task.get("scenario", {})
    behavior = task.get("expected_behavior", {})

    image_id = prov.get("image_id", "unknown")
    tags = _TASK_TYPE_TO_ARE_TAGS.get(coords.get("task_type", ""), [])

    # Build events from turns and expected API calls
    events = []

    # User message events
    for i, turn in enumerate(scenario.get("turns", [])):
        if turn.get("role") == "user":
            events.append({
                "class_name": "Event",
                "event_type": "ENV",
                "event_id": f"msg_{i}",
                "event_time": None,
                "event_relative_time": float(i * 5),
                "dependencies": [f"msg_{i-1}"] if i > 0 else [],
                "action": {
                    "action_id": f"send_{i}",
                    "app": "AgentUserInterface",
                    "function": "send_message_to_agent",
                    "operation_type": None,
                    "args": [
                        {
                            "name": "content",
                            "value": json.dumps(
                                turn.get("content", ""), ensure_ascii=False
                            ),
                            "value_type": "str",
                        }
                    ],
                },
            })

    # Expected API call oracle events
    for j, call in enumerate(behavior.get("api_calls", [])):
        events.append({
            "class_name": "OracleEvent",
            "event_type": "AGENT",
            "event_id": f"api_{j}",
            "event_time": None,
            "event_relative_time": None,
            "event_time_comparator": None,
            "dependencies": ["msg_0"] if events else [],
            "action": {
                "action_id": f"api_call_{j}",
                "app": "GeoApp",
                "function": call.get("method", ""),
                "operation_type": "READ",
                "args": [
                    {"name": k, "value": json.dumps(v), "value_type": "str"}
                    for k, v in call.get("params", {}).items()
                ],
            },
        })

    return {
        "version": "are_simulation_v1",
        "metadata": {
            "definition": {
                "scenario_id": f"geo_{image_id}",
                "seed": None,
                "duration": 120.0,
                "time_increment_in_seconds": 1,
                "start_time": None,
                "run_number": None,
                "hints": [
                    {
                        "hint_type": "TASK_HINT",
                        "content": scenario.get("narrative", ""),
                        "associated_event_id": "msg_0",
                    }
                ],
                "config": None,
                "has_a2a_augmentation": False,
                "has_tool_augmentation": False,
                "has_env_events_augmentation": False,
                "has_exception": False,
                "exception_type": None,
                "exception_message": None,
                "tags": tags,
                "hf_metadata": None,
            },
            "simulation": None,
            "annotation": None,
            "execution": None,
            "runner_config": None,
        },
        "apps": [
            {
                "name": "agent_ui",
                "class_name": "AgentUserInterface",
                "app_state": None,
            },
            {
                "name": "geo",
                "class_name": "GeoApp",
                "app_state": {
                    "pois": prov.get("nearby_pois", []),
                    "location": prov.get("location", {}),
                },
            },
        ],
        "events": events,
        "completed_events": [],
        "world_logs": [],
        "context": None,
        "augmentation": None,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# Main
# ═══════════════════════════════════════════════════════════════════════════════

def main(config: Optional[PipelineConfig] = None) -> Path:
    config = config or PipelineConfig()

    DATA_OUT.mkdir(parents=True, exist_ok=True)
    CARDS_DIR.mkdir(parents=True, exist_ok=True)
    ARE_DIR.mkdir(parents=True, exist_ok=True)

    tasks = read_jsonl(INPUT_PATH)
    if not tasks:
        print("[Module 6] No verified tasks found. Run module 5 first.")
        return XLSX_PATH

    # XLSX export
    xlsx_path = _export_xlsx(tasks)
    print(f"[Module 6] XLSX exported: {xlsx_path}")

    # HTML cards + ARE skeletons
    for task in tasks:
        image_id = task.get("provenance", {}).get("image_id", "unknown")

        # HTML card
        card_path = CARDS_DIR / f"{image_id}.html"
        _export_html_card(task, card_path)

        # ARE skeleton
        skeleton = _export_are_skeleton(task)
        are_path = ARE_DIR / f"{image_id}.json"
        are_path.write_text(
            json.dumps(skeleton, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    print(f"[Module 6] {len(tasks)} HTML cards exported to {CARDS_DIR}")
    print(f"[Module 6] {len(tasks)} ARE skeletons exported to {ARE_DIR}")
    return xlsx_path


if __name__ == "__main__":
    main()
